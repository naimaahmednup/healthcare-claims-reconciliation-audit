"""
Build the Excel audit workbook from the SQL audit outputs.

Design intent: the workbook is not a screenshot of the SQL results. It reloads
the reconciliation grain and re-derives the payment checks in native Excel
formulas, then cross-foots its own answers against what PostgreSQL found. If the
two disagree on a single claim, the Dashboard says so.

Run after scripts/run_audit.sh:
    python scripts/build_workbook.py
"""

import csv
import os
from datetime import date

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.utils import get_column_letter

OUT = "output"
XLSX = os.path.join(OUT, "Claims_Reconciliation_Audit.xlsx")

# ---------------------------------------------------------------------------
# House style
# ---------------------------------------------------------------------------
FONT = "Arial"
INK        = "1F2933"
MUTED      = "5B6770"
RULE       = "D5DAE0"
BAND       = "F4F6F8"
HEADER_BG  = "0D366B"     # blue 700
ACCENT     = "256ABF"     # blue 500, the single series hue
INPUT_BLUE = "0000FF"     # hardcoded input convention
ASSUMPTION = "FFF3C4"
GOOD       = "0CA30C"
WARNING    = "FAB219"
CRITICAL   = "D03B3B"
GOOD_BG    = "E3F5E3"
BAD_BG     = "FBE4E4"
WARN_BG    = "FEF3DA"

MONEY = '$#,##0.00;($#,##0.00);"-"'
MONEY0 = '$#,##0;($#,##0);"-"'
INT = '#,##0'
PCT = '0.000%'
PCT1 = '0.0%'

thin = Side(style="thin", color=RULE)
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)
UNDER = Border(bottom=Side(style="thin", color=RULE))


def read_csv(name):
    with open(os.path.join(OUT, name), newline="") as fh:
        return list(csv.DictReader(fh))


def num(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except ValueError:
        return v


def as_date(v):
    if not v:
        return None
    try:
        return date.fromisoformat(v)
    except ValueError:
        return None


def style_header(ws, row, ncols, start=1):
    for c in range(start, start + ncols):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = BOX
    ws.row_dimensions[row].height = 30


def title_block(ws, title, subtitle, width=8):
    ws["A1"] = title
    ws["A1"].font = Font(name=FONT, size=16, bold=True, color=INK)
    ws["A2"] = subtitle
    ws["A2"].font = Font(name=FONT, size=10, color=MUTED)
    ws.row_dimensions[1].height = 22


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------------------------------------------------------------------------
# Load the audit outputs
# ---------------------------------------------------------------------------
summary   = read_csv("dq_run_summary.csv")
register  = read_csv("exception_register.csv")
recon     = read_csv("claim_reconciliation.csv")
payers    = read_csv("payer_scorecard.csv")
validation = read_csv("suite_validation.csv")
controls  = read_csv("control_totals.csv")[0]

# Row counts of the extracts, straight from the landing tables. These are the
# only figures in the workbook that Excel cannot recompute for itself, so they
# are entered as documented inputs on the Read Me tab.
CLAIM_ROWS_RECEIVED = 25230
REMIT_ROWS_RECEIVED = 23649

QUEUE = {
    "DQ-01": "Interface / EDI",
    "DQ-02": "Charge Entry",
    "DQ-03": "Cash Posting",
    "DQ-04": "Cash Posting",
    "DQ-05": "A/R Follow-up",
    "DQ-06": "Payment Variance",
    "DQ-07": "Refunds / Credit Balance",
    "DQ-08": "Contract Management",
    "DQ-09": "Payer Maintenance",
    "DQ-10": "Provider Data Management",
    "DQ-11": "Source System Owner",
    "DQ-12": "Charge Entry",
}

ROOT_CAUSE = {
    "DQ-01": "Interface engine re-sent a batch without collapsing on the claim key.",
    "DQ-02": "Charge entry re-keyed an encounter that was already billed; no duplicate guard on the patient/DOS/CPT combination.",
    "DQ-03": "The same 835 file was posted twice; the posting job has no idempotency key on check number plus claim.",
    "DQ-04": "Payer remitted against a claim number the billing system never issued - likely a payer-side claim number or a claim purged after submission.",
    "DQ-05": "Claim status was advanced to PAID manually, or the 835 was never ingested for that check.",
    "DQ-06": "Partial adjustments posted without the matching contractual write-off, so the claim no longer foots to the charge.",
    "DQ-07": "Payer paid above the billed charge - usually a duplicate payer-side adjudication or a coordination-of-benefits error.",
    "DQ-08": "Payer priced below the loaded contract rate; either the fee schedule is stale on their side or ours.",
    "DQ-09": "New payer or plan not added to the payer master before claims began flowing.",
    "DQ-10": "Provider record created without NPI validation at entry.",
    "DQ-11": "Date mapping defect in the source feed; dates are being written without timezone or sequence validation.",
    "DQ-12": "Required fields not enforced at charge entry; claims can be saved incomplete.",
}

SLA = {
    "CRITICAL": "24 hours",
    "HIGH": "3 business days",
    "MEDIUM": "5 business days",
}

wb = Workbook()

REG_ROWS = len(register)
REG_LAST = REG_ROWS + 1          # last data row on the register sheet
RECON_ROWS = len(recon)
RECON_LAST = RECON_ROWS + 1

# Contiguous row spans per check on the register (it is exported ordered by
# check_id), so the Recon Detail lookups scan only the rows they need instead of
# the whole register 25,000 times over.
spans = {}
for i, r in enumerate(register, start=2):
    cid = r["check_id"]
    if cid not in spans:
        spans[cid] = [i, i]
    spans[cid][1] = i

REG = "'Exception Register'"
RECON = "'Recon Detail'"
READ = "'Read Me'"


# ===========================================================================
# 1. Read Me
# ===========================================================================
ws = wb.active
ws.title = "Read Me"
set_widths(ws, [46, 3, 34, 3, 62])
title_block(ws, "Healthcare Claims Reconciliation & Data Quality Audit",
            "Claims source data reconciled against payer remittance, with a 12-check data quality suite. "
            "Synthetic data - no real patient or payer information.")


def section(row, text):
    ws.cell(row=row, column=1, value=text).font = Font(name=FONT, size=11, bold=True, color=HEADER_BG)
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = UNDER


def param(row, label, value, fmt=None, kind="formula", note=None):
    ws.cell(row=row, column=1, value=label).font = Font(name=FONT, size=10, color=INK)
    c = ws.cell(row=row, column=3, value=value)
    colour = INPUT_BLUE if kind in ("input", "assumption") else INK
    c.font = Font(name=FONT, size=10, bold=True, color=colour)
    if kind == "assumption":
        c.fill = PatternFill("solid", fgColor=ASSUMPTION)
    if fmt:
        c.number_format = fmt
    c.alignment = Alignment(horizontal="right")
    if note:
        n = ws.cell(row=row, column=5, value=note)
        n.font = Font(name=FONT, size=9, italic=True, color=MUTED)


section(4, "RUN PARAMETERS")
param(5, "Audit run identifier", "RUN-2026-08-29", kind="input",
      note="Stamped on every row of the exception register.")
param(6, "Audit reference date", date(2026, 8, 29), fmt="yyyy-mm-dd", kind="input",
      note="Used by DQ-11 to decide what counts as a future date of service.")
param(7, "Database engine", "PostgreSQL 16", kind="input")
param(8, "Claims extract", "data/raw/claims_source.csv", kind="input")
param(9, "Remittance extract", "data/raw/payments_remittance.csv", kind="input")
param(10, "Claim rows received in the extract", CLAIM_ROWS_RECEIVED, fmt=INT, kind="input",
      note="Source: select count(*) from rcm.stg_claims_source. Includes duplicate rows, which is the point of DQ-01.")
param(11, "Remittance rows received in the extract", REMIT_ROWS_RECEIVED, fmt=INT, kind="input",
      note="Source: select count(*) from rcm.stg_payments_remittance.")
param(12, "Unique claims after de-duplication",
      f"=COUNTA({RECON}!$A$2:$A${RECON_LAST})", fmt=INT,
      note="Counted live from the Recon Detail tab - the reconciliation denominator.")
param(13, "Total rows compared across both systems", "=C10+C11", fmt=INT)
param(14, "Balance tolerance (USD)", 0.01, fmt=MONEY, kind="assumption",
      note="A claim foots if |billed - (paid + patient responsibility + adjustment)| is within this. Rounding only.")
param(15, "Underpayment threshold", 0.10, fmt=PCT1, kind="assumption",
      note="DQ-08 fires when the allowed amount falls more than this below the contracted rate.")

section(17, "HOW THIS WORKBOOK IS BUILT")
notes = [
    "The audit runs in PostgreSQL. Twelve checks, one per failure condition, are defined in sql/checks/ and land in a single",
    "exception register. This workbook then reloads the reconciliation grain and re-derives the three payment checks",
    "(DQ-05, DQ-06, DQ-07) in native Excel formulas, independently of the SQL.",
    "",
    "The Dashboard cross-foots the two. If Excel and SQL disagree about a single claim out of 25,140, the count of",
    "disagreements on the Dashboard stops being zero. That is the control - not the fact that the queries ran.",
    "",
    "Formulas use INDEX/MATCH, SUMIF, COUNTIF and SUMPRODUCT rather than XLOOKUP or FILTER, so the workbook opens and",
    "recalculates correctly in Excel 2016, LibreOffice and Google Sheets as well as Microsoft 365.",
]
for i, line in enumerate(notes):
    c = ws.cell(row=18 + i, column=1, value=line)
    c.font = Font(name=FONT, size=10, color=INK if line else MUTED)

section(28, "TAB GUIDE")
guide = [
    ("Dashboard", "Control totals, the SQL/Excel cross-foot, exceptions by severity and by check."),
    ("Check Summary", "The twelve checks: what each one tests, what it found, and what it is worth in dollars."),
    ("Defect Log", "One entry per defect class: probable root cause, owning queue, exposure, SLA and status."),
    ("Exception Register", "Every exception raised, one row each, with the evidence and the queue it was routed to."),
    ("Recon Detail", "The reconciliation grain - one row per unique claim, with the Excel-side checks recomputed live."),
    ("Payer Scorecard", "Where the exceptions concentrate, normalised per 1,000 claims so volume does not mislead."),
    ("Suite Validation", "Recall and precision of each check, scored against the ledger of seeded defects."),
]
ws.cell(row=29, column=1, value="Tab").font = Font(name=FONT, size=10, bold=True)
ws.cell(row=29, column=3, value="What it holds").font = Font(name=FONT, size=10, bold=True)
for i, (tab, desc) in enumerate(guide):
    ws.cell(row=30 + i, column=1, value=tab).font = Font(name=FONT, size=10, bold=True, color=ACCENT)
    c = ws.cell(row=30 + i, column=3, value=desc)
    c.font = Font(name=FONT, size=10, color=INK)
    ws.merge_cells(start_row=30 + i, start_column=3, end_row=30 + i, end_column=5)

section(39, "CONVENTIONS")
conv = [
    ("Blue text", "A hardcoded input taken from the SQL run. Every one names its source in the note beside it."),
    ("Yellow fill", "A threshold you may want to change. Formulas reference the cell, never the number."),
    ("Black text", "Calculated by a formula in this workbook."),
    ("Red / amber / green fill", "Status, always paired with the word - never colour on its own."),
]
for i, (k, v) in enumerate(conv):
    ws.cell(row=40 + i, column=1, value=k).font = Font(name=FONT, size=10, bold=True)
    c = ws.cell(row=40 + i, column=3, value=v)
    c.font = Font(name=FONT, size=10, color=INK)
    ws.merge_cells(start_row=40 + i, start_column=3, end_row=40 + i, end_column=5)

ws.cell(row=45, column=1,
        value="Data note: all claims, payments, patients, providers and NPIs are generated. "
              "Payer names are real organisations used only to make the payer mix realistic; "
              "no real contract terms, rates or claim data are represented.").font = \
    Font(name=FONT, size=9, italic=True, color=MUTED)
ws.merge_cells(start_row=45, start_column=1, end_row=45, end_column=5)
ws.sheet_view.showGridLines = False

# ===========================================================================
# 2. Check Summary
# ===========================================================================
cs = wb.create_sheet("Check Summary")
title_block(cs, "Data Quality Check Summary",
            "One row per check. Counts and dollar values are formulas over the Exception Register, "
            "so this tab re-foots itself if the register changes.")
cs_head = ["Check", "Check name", "Dimension", "Severity", "Owning queue", "Exceptions",
           "Financial exposure (USD)", "Exception rate", "Result", "Failure condition", "Escalation step"]
for i, h in enumerate(cs_head, start=1):
    cs.cell(row=4, column=i, value=h)
style_header(cs, 4, len(cs_head))
set_widths(cs, [9, 34, 14, 11, 24, 12, 20, 13, 10, 74, 78])

for i, row in enumerate(summary):
    r = 5 + i
    cid = row["check_id"]
    cs.cell(row=r, column=1, value=cid)
    cs.cell(row=r, column=2, value=row["check_name"])
    cs.cell(row=r, column=3, value=row["dimension"])
    cs.cell(row=r, column=4, value=row["severity"])
    cs.cell(row=r, column=5, value=QUEUE[cid])
    cs.cell(row=r, column=6, value=f"=COUNTIF({REG}!$B$2:$B${REG_LAST},$A{r})").number_format = INT
    cs.cell(row=r, column=7,
            value=f"=SUMIF({REG}!$B$2:$B${REG_LAST},$A{r},{REG}!$M$2:$M${REG_LAST})").number_format = MONEY
    cs.cell(row=r, column=8, value=f"=IF({READ}!$C$12=0,0,$F{r}/{READ}!$C$12)").number_format = PCT
    cs.cell(row=r, column=9, value=f'=IF($F{r}=0,"PASS","FAIL")')
    cs.cell(row=r, column=10, value=row["failure_condition"])
    cs.cell(row=r, column=11, value=row["escalation"])
    for c in range(1, 12):
        cell = cs.cell(row=r, column=c)
        cell.font = Font(name=FONT, size=10, color=INK)
        cell.border = BOX
        cell.alignment = Alignment(vertical="top",
                                   wrap_text=c in (2, 10, 11),
                                   horizontal="left" if c not in (6, 7, 8) else "right")
    if i % 2 == 1:
        for c in range(1, 12):
            cs.cell(row=r, column=c).fill = PatternFill("solid", fgColor=BAND)

LAST_CS = 4 + len(summary)
cs.cell(row=LAST_CS + 1, column=5, value="TOTAL").font = Font(name=FONT, size=10, bold=True)
cs.cell(row=LAST_CS + 1, column=6, value=f"=SUM(F5:F{LAST_CS})").number_format = INT
cs.cell(row=LAST_CS + 1, column=7, value=f"=SUM(G5:G{LAST_CS})").number_format = MONEY
for c in (5, 6, 7):
    cell = cs.cell(row=LAST_CS + 1, column=c)
    cell.font = Font(name=FONT, size=10, bold=True, color=INK)
    cell.border = Border(top=Side(style="double", color=INK))

cs.conditional_formatting.add(f"I5:I{LAST_CS}",
    CellIsRule(operator="equal", formula=['"FAIL"'],
               fill=PatternFill("solid", fgColor=BAD_BG),
               font=Font(name=FONT, size=10, bold=True, color=CRITICAL)))
cs.conditional_formatting.add(f"I5:I{LAST_CS}",
    CellIsRule(operator="equal", formula=['"PASS"'],
               fill=PatternFill("solid", fgColor=GOOD_BG),
               font=Font(name=FONT, size=10, bold=True, color=GOOD)))
cs.conditional_formatting.add(f"D5:D{LAST_CS}",
    CellIsRule(operator="equal", formula=['"CRITICAL"'],
               font=Font(name=FONT, size=10, bold=True, color=CRITICAL)))
cs.conditional_formatting.add(f"D5:D{LAST_CS}",
    CellIsRule(operator="equal", formula=['"HIGH"'],
               font=Font(name=FONT, size=10, bold=True, color="B26A00")))
cs.conditional_formatting.add(f"F5:F{LAST_CS}",
    ColorScaleRule(start_type="min", start_color="FFFFFF", end_type="max", end_color="9EC5F4"))
cs.freeze_panes = "B5"
cs.auto_filter.ref = f"A4:K{LAST_CS}"
cs.sheet_view.showGridLines = False

# ===========================================================================
# 3. Defect Log
# ===========================================================================
dl = wb.create_sheet("Defect Log")
title_block(dl, "Defect Log",
            "One entry per defect class raised by the 2026-08-29 run. Volume and exposure are formulas "
            "over the Exception Register; root cause is the analyst's assessment, not a system output.")
dl_head = ["Defect ID", "Raised", "Check", "Defect title", "Severity", "Probable root cause",
           "Volume", "Financial exposure (USD)", "Owning queue", "Target SLA", "Status", "Evidence file"]
for i, h in enumerate(dl_head, start=1):
    dl.cell(row=4, column=i, value=h)
style_header(dl, 4, len(dl_head))
set_widths(dl, [11, 12, 9, 34, 11, 84, 10, 20, 24, 16, 10, 46])

EVIDENCE = {
    "DQ-01": "output/exceptions/DQ-01_duplicate_claim_ids.csv",
    "DQ-02": "output/exceptions/DQ-02_duplicate_billing_fingerprint.csv",
    "DQ-03": "output/exceptions/DQ-03_duplicate_remittance_posting.csv",
    "DQ-04": "output/exceptions/DQ-04_orphaned_remittance.csv",
    "DQ-05": "output/exceptions/DQ-05_missing_remittance.csv",
    "DQ-06": "output/exceptions/DQ-06_claim_payment_mismatch.csv",
    "DQ-07": "output/exceptions/DQ-07_overpayment.csv",
    "DQ-08": "output/exceptions/DQ-08_allowed_below_contract.csv",
    "DQ-09": "output/exceptions/DQ-09_unmapped_payer.csv",
    "DQ-10": "output/exceptions/DQ-10_invalid_npi.csv",
    "DQ-11": "output/exceptions/DQ-11_temporal_integrity.csv",
    "DQ-12": "output/exceptions/DQ-12_mandatory_fields.csv",
}

for i, row in enumerate(summary):
    r = 5 + i
    cid = row["check_id"]
    dl.cell(row=r, column=1, value=f"DEF-{i + 1:03d}")
    dl.cell(row=r, column=2, value=date(2026, 8, 29)).number_format = "yyyy-mm-dd"
    dl.cell(row=r, column=3, value=cid)
    dl.cell(row=r, column=4, value=row["check_name"])
    dl.cell(row=r, column=5, value=row["severity"])
    dl.cell(row=r, column=6, value=ROOT_CAUSE[cid])
    dl.cell(row=r, column=7, value=f"=COUNTIF({REG}!$B$2:$B${REG_LAST},$C{r})").number_format = INT
    dl.cell(row=r, column=8,
            value=f"=SUMIF({REG}!$B$2:$B${REG_LAST},$C{r},{REG}!$M$2:$M${REG_LAST})").number_format = MONEY
    dl.cell(row=r, column=9, value=QUEUE[cid])
    dl.cell(row=r, column=10, value=SLA[row["severity"]])
    dl.cell(row=r, column=11, value="OPEN")
    dl.cell(row=r, column=12, value=EVIDENCE[cid])
    for c in range(1, 13):
        cell = dl.cell(row=r, column=c)
        cell.font = Font(name=FONT, size=10, color=INK)
        cell.border = BOX
        cell.alignment = Alignment(vertical="top", wrap_text=c in (4, 6, 12),
                                   horizontal="right" if c in (7, 8) else "left")
    if i % 2 == 1:
        for c in range(1, 13):
            dl.cell(row=r, column=c).fill = PatternFill("solid", fgColor=BAND)

LAST_DL = 4 + len(summary)
dl.conditional_formatting.add(f"E5:E{LAST_DL}",
    CellIsRule(operator="equal", formula=['"CRITICAL"'],
               fill=PatternFill("solid", fgColor=BAD_BG),
               font=Font(name=FONT, size=10, bold=True, color=CRITICAL)))
dl.conditional_formatting.add(f"E5:E{LAST_DL}",
    CellIsRule(operator="equal", formula=['"HIGH"'],
               fill=PatternFill("solid", fgColor=WARN_BG),
               font=Font(name=FONT, size=10, bold=True, color="B26A00")))
dl.freeze_panes = "D5"
dl.auto_filter.ref = f"A4:L{LAST_DL}"
dl.sheet_view.showGridLines = False

# ===========================================================================
# 4. Exception Register   (header on row 1, data from row 2 - the lookup
#    formulas on other tabs address these rows directly)
# ===========================================================================
er = wb.create_sheet("Exception Register")
er_head = ["Run ID", "Check", "Severity", "Entity type", "Entity key", "Claim ID", "Payer ID",
           "Payer name", "Date of service", "CPT", "Billed amount", "Amount impact",
           "Absolute impact", "Assigned queue", "Disposition", "Finding"]
for i, h in enumerate(er_head, start=1):
    er.cell(row=1, column=i, value=h)
style_header(er, 1, len(er_head))
set_widths(er, [16, 9, 11, 12, 18, 18, 14, 26, 15, 8, 14, 15, 15, 26, 12, 96])

for i, row in enumerate(register):
    r = 2 + i
    er.cell(row=r, column=1, value=row["run_id"])
    er.cell(row=r, column=2, value=row["check_id"])
    er.cell(row=r, column=3, value=row["severity"])
    er.cell(row=r, column=4, value=row["entity_type"])
    er.cell(row=r, column=5, value=row["entity_key"])
    er.cell(row=r, column=6, value=row["claim_id"] or None)
    er.cell(row=r, column=7, value=row["payer_id"] or None)
    er.cell(row=r, column=8, value=row["payer_name"] or None)
    er.cell(row=r, column=9, value=as_date(row["date_of_service"])).number_format = "yyyy-mm-dd"
    er.cell(row=r, column=10, value=row["cpt_code"] or None)
    er.cell(row=r, column=11, value=num(row["billed_amount"])).number_format = MONEY
    er.cell(row=r, column=12, value=num(row["amount_impact"])).number_format = MONEY
    er.cell(row=r, column=13, value=f"=ABS(L{r})").number_format = MONEY
    er.cell(row=r, column=14, value=row["assigned_queue"])
    er.cell(row=r, column=15, value=row["disposition"])
    er.cell(row=r, column=16, value=row["finding"])
    for c in range(1, 17):
        er.cell(row=r, column=c).font = Font(name=FONT, size=10, color=INK)

er.conditional_formatting.add(f"C2:C{REG_LAST}",
    CellIsRule(operator="equal", formula=['"CRITICAL"'],
               font=Font(name=FONT, size=10, bold=True, color=CRITICAL)))
er.conditional_formatting.add(f"C2:C{REG_LAST}",
    CellIsRule(operator="equal", formula=['"HIGH"'],
               font=Font(name=FONT, size=10, bold=True, color="B26A00")))
er.conditional_formatting.add(f"M2:M{REG_LAST}",
    ColorScaleRule(start_type="min", start_color="FFFFFF", end_type="max", end_color="86B6EF"))
er.freeze_panes = "C2"
er.auto_filter.ref = f"A1:P{REG_LAST}"

# ===========================================================================
# 5. Recon Detail   (header row 1, data from row 2)
#
# Columns A-P are the reconciliation grain as loaded. Q-U are the Excel side of
# the audit: the workbook re-derives the balance identity and the three payment
# checks from the raw amounts, then compares its own verdict to the SQL one.
# ===========================================================================
rd = wb.create_sheet("Recon Detail")
rd_head = ["Claim ID", "Patient ID", "Provider NPI", "Payer ID", "Payer name", "Date of service",
           "Submit date", "CPT", "Claim status", "Billed amount", "Expected allowed",
           "Remit lines", "Allowed (remit)", "Paid (remit)", "Patient resp (remit)",
           "Adjustment (remit)",
           "Accounted (Excel)", "Variance (Excel)", "Excel verdict", "SQL check", "Excel vs SQL"]
for i, h in enumerate(rd_head, start=1):
    rd.cell(row=1, column=i, value=h)
style_header(rd, 1, len(rd_head))
set_widths(rd, [18, 12, 14, 14, 26, 15, 13, 8, 13, 14, 16, 11, 15, 14, 17, 17,
                17, 16, 22, 11, 13])

FIELDS = ["claim_id", "patient_id", "provider_npi", "payer_id", "payer_name",
          "date_of_service", "claim_submit_date", "cpt_code", "claim_status",
          "billed_amount", "expected_allowed_amount", "remit_line_count",
          "allowed_total", "paid_total", "patient_resp_total", "adjustment_total"]

TOL = f"{READ}!$C$14"


def span_ref(cid):
    lo, hi = spans[cid]
    return f"{REG}!$F${lo}:$F${hi}"


for i, row in enumerate(recon):
    r = 2 + i
    for c, f in enumerate(FIELDS, start=1):
        v = row[f]
        if f in ("date_of_service", "claim_submit_date"):
            cell = rd.cell(row=r, column=c, value=as_date(v))
            cell.number_format = "yyyy-mm-dd"
        elif f in ("billed_amount", "expected_allowed_amount", "allowed_total",
                   "paid_total", "patient_resp_total", "adjustment_total"):
            cell = rd.cell(row=r, column=c, value=num(v))
            cell.number_format = MONEY
        elif f == "remit_line_count":
            cell = rd.cell(row=r, column=c, value=int(v or 0))
        else:
            cell = rd.cell(row=r, column=c, value=v or None)

    rd.cell(row=r, column=17, value=f"=SUM(N{r},O{r},P{r})").number_format = MONEY
    rd.cell(row=r, column=18, value=f"=J{r}-Q{r}").number_format = MONEY
    # The Excel verdict follows exactly the same precedence the SQL suite uses,
    # so the two are comparable claim by claim.
    rd.cell(row=r, column=19, value=(
        f'=IF(OR(J{r}="",J{r}<=0),"INVALID CHARGE",'
        f'IF(L{r}=0,IF(I{r}="PAID","NO REMITTANCE (PAID)","NO REMITTANCE"),'
        f'IF(L{r}>1,"DUPLICATE POSTING",'
        f'IF(N{r}>J{r},"OVERPAYMENT",'
        f'IF(ABS(R{r})>{TOL},"MISMATCH","BALANCED")))))'))
    rd.cell(row=r, column=20, value=(
        f'=IF(COUNTIF({span_ref("DQ-05")},$A{r})>0,"DQ-05",'
        f'IF(COUNTIF({span_ref("DQ-06")},$A{r})>0,"DQ-06",'
        f'IF(COUNTIF({span_ref("DQ-07")},$A{r})>0,"DQ-07","")))'))
    rd.cell(row=r, column=21, value=(
        f'=IF(OR('
        f'AND(S{r}="BALANCED",T{r}=""),'
        f'AND(S{r}="INVALID CHARGE",T{r}=""),'
        f'AND(S{r}="NO REMITTANCE",T{r}=""),'
        f'AND(S{r}="NO REMITTANCE (PAID)",T{r}="DQ-05"),'
        f'AND(S{r}="DUPLICATE POSTING",T{r}=""),'
        f'AND(S{r}="MISMATCH",T{r}="DQ-06"),'
        f'AND(S{r}="OVERPAYMENT",T{r}="DQ-07")'
        f'),"AGREE","CHECK")'))
    for c in range(1, 22):
        rd.cell(row=r, column=c).font = Font(name=FONT, size=10, color=INK)

rd.conditional_formatting.add(f"S2:S{RECON_LAST}",
    CellIsRule(operator="equal", formula=['"BALANCED"'],
               font=Font(name=FONT, size=10, color=GOOD)))
for verdict in ("MISMATCH", "OVERPAYMENT", "DUPLICATE POSTING", "INVALID CHARGE"):
    rd.conditional_formatting.add(f"S2:S{RECON_LAST}",
        CellIsRule(operator="equal", formula=[f'"{verdict}"'],
                   fill=PatternFill("solid", fgColor=BAD_BG),
                   font=Font(name=FONT, size=10, bold=True, color=CRITICAL)))
rd.conditional_formatting.add(f"U2:U{RECON_LAST}",
    CellIsRule(operator="equal", formula=['"CHECK"'],
               fill=PatternFill("solid", fgColor=BAD_BG),
               font=Font(name=FONT, size=10, bold=True, color=CRITICAL)))
rd.conditional_formatting.add(f"R2:R{RECON_LAST}",
    FormulaRule(formula=[f"ABS(R2)>{TOL}"],
                font=Font(name=FONT, size=10, bold=True, color=CRITICAL)))
rd.freeze_panes = "B2"
rd.auto_filter.ref = f"A1:U{RECON_LAST}"

# ===========================================================================
# 6. Payer Scorecard
# ===========================================================================
ps = wb.create_sheet("Payer Scorecard")
title_block(ps, "Payer Scorecard",
            "Exception concentration by payer. Raw volume follows claim volume, so the last column "
            "normalises to exceptions per 1,000 claims - that is the column to read.")
ps_head = ["Payer ID", "Payer name", "Claims", "Billed (USD)", "Exceptions",
           "Financial exposure (USD)", "Exceptions per 1,000 claims"]
for i, h in enumerate(ps_head, start=1):
    ps.cell(row=4, column=i, value=h)
style_header(ps, 4, len(ps_head))
set_widths(ps, [16, 32, 12, 18, 13, 24, 26])

RD_PAYER = f"{RECON}!$D$2:$D${RECON_LAST}"
RD_BILL = f"{RECON}!$J$2:$J${RECON_LAST}"
RG_PAYER = f"{REG}!$G$2:$G${REG_LAST}"
RG_IMP = f"{REG}!$M$2:$M${REG_LAST}"

for i, row in enumerate(payers):
    r = 5 + i
    pid = row["payer_id"]
    unmapped = (pid == "UNMAPPED")
    ps.cell(row=r, column=1, value=pid)
    ps.cell(row=r, column=2, value=row["payer_name"])
    if unmapped:
        # The payer id is null on these claims, so they are matched on blankness
        # rather than on the literal label used in the SQL rollup.
        ps.cell(row=r, column=3, value=f"=COUNTBLANK({RD_PAYER})")
        ps.cell(row=r, column=4, value=f"=SUMPRODUCT(--({RD_PAYER}=\"\"),{RD_BILL})")
        ps.cell(row=r, column=5, value=f"=COUNTBLANK({RG_PAYER})")
        ps.cell(row=r, column=6, value=f"=SUMPRODUCT(--({RG_PAYER}=\"\"),{RG_IMP})")
    else:
        ps.cell(row=r, column=3, value=f"=COUNTIF({RD_PAYER},$A{r})")
        ps.cell(row=r, column=4, value=f"=SUMIF({RD_PAYER},$A{r},{RD_BILL})")
        ps.cell(row=r, column=5, value=f"=COUNTIF({RG_PAYER},$A{r})")
        ps.cell(row=r, column=6, value=f"=SUMIF({RG_PAYER},$A{r},{RG_IMP})")
    ps.cell(row=r, column=7, value=f"=IF($C{r}=0,0,1000*$E{r}/$C{r})").number_format = '#,##0.0'
    ps.cell(row=r, column=3).number_format = INT
    ps.cell(row=r, column=4).number_format = MONEY0
    ps.cell(row=r, column=5).number_format = INT
    ps.cell(row=r, column=6).number_format = MONEY
    for c in range(1, 8):
        cell = ps.cell(row=r, column=c)
        cell.font = Font(name=FONT, size=10, color=INK)
        cell.border = BOX
    if i % 2 == 1:
        for c in range(1, 8):
            ps.cell(row=r, column=c).fill = PatternFill("solid", fgColor=BAND)

LAST_PS = 4 + len(payers)
ps.cell(row=LAST_PS + 1, column=2, value="TOTAL").font = Font(name=FONT, size=10, bold=True)
for c, fmt in ((3, INT), (4, MONEY0), (5, INT), (6, MONEY)):
    cell = ps.cell(row=LAST_PS + 1, column=c,
                   value=f"=SUM({get_column_letter(c)}5:{get_column_letter(c)}{LAST_PS})")
    cell.number_format = fmt
    cell.font = Font(name=FONT, size=10, bold=True, color=INK)
    cell.border = Border(top=Side(style="double", color=INK))
ps.conditional_formatting.add(f"G5:G{LAST_PS}",
    ColorScaleRule(start_type="min", start_color="FFFFFF", end_type="max", end_color="86B6EF"))
ps.freeze_panes = "C5"
ps.sheet_view.showGridLines = False

# ===========================================================================
# 7. Suite Validation
# ===========================================================================
sv = wb.create_sheet("Suite Validation")
title_block(sv, "Audit Suite Validation",
            "Every check scored against the ledger of defects seeded into the dataset. Recall answers "
            "'did it find them all', precision answers 'did it flag anything it should not have'.")
sv_head = ["Check", "Check name", "Defects seeded", "Exceptions detected", "True positives",
           "False negatives", "False positives", "Recall", "Precision", "Result"]
for i, h in enumerate(sv_head, start=1):
    sv.cell(row=4, column=i, value=h)
style_header(sv, 4, len(sv_head))
set_widths(sv, [9, 34, 15, 20, 14, 15, 14, 11, 11, 11])

for i, row in enumerate(validation):
    r = 5 + i
    sv.cell(row=r, column=1, value=row["check_id"])
    sv.cell(row=r, column=2, value=row["check_name"])
    c = sv.cell(row=r, column=3, value=int(row["injected_defects"]))
    c.font = Font(name=FONT, size=10, bold=True, color=INPUT_BLUE)
    sv.cell(row=r, column=4, value=f"=COUNTIF({REG}!$B$2:$B${REG_LAST},$A{r})")
    c = sv.cell(row=r, column=5, value=int(row["true_positives"]))
    c.font = Font(name=FONT, size=10, bold=True, color=INPUT_BLUE)
    sv.cell(row=r, column=6, value=f"=$C{r}-$E{r}")
    sv.cell(row=r, column=7, value=f"=$D{r}-$E{r}")
    sv.cell(row=r, column=8, value=f"=IF($C{r}=0,\"\",$E{r}/$C{r})").number_format = PCT1
    sv.cell(row=r, column=9, value=f"=IF($D{r}=0,\"\",$E{r}/$D{r})").number_format = PCT1
    sv.cell(row=r, column=10, value=f'=IF(AND($F{r}=0,$G{r}=0),"PASS","REVIEW")')
    for c_ in range(1, 11):
        cell = sv.cell(row=r, column=c_)
        if cell.font.color is None or cell.font.color.rgb != f"00{INPUT_BLUE}":
            cell.font = Font(name=FONT, size=10, color=INK)
        cell.border = BOX
    for c_ in (3, 4, 5, 6, 7):
        sv.cell(row=r, column=c_).number_format = INT

LAST_SV = 4 + len(validation)
sv.conditional_formatting.add(f"J5:J{LAST_SV}",
    CellIsRule(operator="equal", formula=['"PASS"'],
               fill=PatternFill("solid", fgColor=GOOD_BG),
               font=Font(name=FONT, size=10, bold=True, color=GOOD)))
sv.conditional_formatting.add(f"J5:J{LAST_SV}",
    CellIsRule(operator="equal", formula=['"REVIEW"'],
               fill=PatternFill("solid", fgColor=WARN_BG),
               font=Font(name=FONT, size=10, bold=True, color="B26A00")))
sv.cell(row=LAST_SV + 2, column=1,
        value="Defects seeded and true positives come from the SQL join between the exception register and "
              "data/raw/_injected_defect_ledger.csv (view rcm.v_suite_validation). Excel cannot recompute a "
              "join it does not hold, so they are entered as documented inputs; everything else on this tab is a formula.")
sv.cell(row=LAST_SV + 2, column=1).font = Font(name=FONT, size=9, italic=True, color=MUTED)
sv.merge_cells(start_row=LAST_SV + 2, start_column=1, end_row=LAST_SV + 2, end_column=10)
sv.cell(row=LAST_SV + 4, column=1,
        value="Negative control: the same 25,000 claims regenerated with no defects seeded raise zero exceptions "
              "across all twelve checks (tests/negative_control.sh).")
sv.cell(row=LAST_SV + 4, column=1).font = Font(name=FONT, size=9, italic=True, color=MUTED)
sv.merge_cells(start_row=LAST_SV + 4, start_column=1, end_row=LAST_SV + 4, end_column=10)
sv.freeze_panes = "C5"
sv.sheet_view.showGridLines = False

# ===========================================================================
# 8. Dashboard  (created last so it can reference every other tab, then moved
#    to the front)
# ===========================================================================
db = wb.create_sheet("Dashboard")
title_block(db, "Claims Reconciliation - Audit Dashboard",
            "Run RUN-2026-08-29  |  PostgreSQL 16  |  every figure below is a formula over the tabs in this workbook.")
set_widths(db, [26, 16, 26, 16, 26, 16, 26, 16])

CS_FIRST, CS_LAST = 5, LAST_CS


def tile(row, col, label, formula, fmt, tone=None):
    lc = db.cell(row=row, column=col, value=label)
    lc.font = Font(name=FONT, size=9, bold=True, color=MUTED)
    lc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    db.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
    vc = db.cell(row=row + 1, column=col, value=formula)
    vc.font = Font(name=FONT, size=18, bold=True, color=tone or HEADER_BG)
    vc.number_format = fmt
    vc.alignment = Alignment(horizontal="left", vertical="center")
    db.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 1)
    for rr in (row, row + 1):
        for cc in (col, col + 1):
            db.cell(row=rr, column=cc).fill = PatternFill("solid", fgColor=BAND)
    db.row_dimensions[row].height = 26
    db.row_dimensions[row + 1].height = 30


tile(4, 1, "UNIQUE CLAIMS REVIEWED", f"={READ}!$C$12", INT)
tile(4, 3, "ROWS COMPARED ACROSS BOTH SYSTEMS", f"={READ}!$C$13", INT)
tile(4, 5, "EXCEPTIONS RAISED", f"=COUNTA({REG}!$A$2:$A${REG_LAST})", INT, CRITICAL)
tile(4, 7, "EXCEPTION RATE PER CLAIM",
     f"=IF({READ}!$C$12=0,0,COUNTA({REG}!$A$2:$A${REG_LAST})/{READ}!$C$12)", PCT1, CRITICAL)

tile(7, 1, "CRITICAL-SEVERITY EXCEPTIONS",
     f'=COUNTIF({REG}!$C$2:$C${REG_LAST},"CRITICAL")', INT, CRITICAL)
tile(7, 3, "FINANCIAL EXPOSURE FLAGGED",
     f"=SUM('Check Summary'!$G${CS_FIRST}:$G${CS_LAST})", MONEY0, CRITICAL)
tile(7, 5, "BILLED VALUE IN SCOPE", f"=SUM({RECON}!$J$2:$J${RECON_LAST})", MONEY0)
tile(7, 7, "NET VARIANCE, ADJUDICATED CLAIMS",
     f'=SUMIF({RECON}!$L$2:$L${RECON_LAST},">0",{RECON}!$R$2:$R${RECON_LAST})', MONEY0, CRITICAL)


def section_head(row, text, span=8):
    c = db.cell(row=row, column=1, value=text)
    c.font = Font(name=FONT, size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=HEADER_BG)
    c.alignment = Alignment(horizontal="left", vertical="center")
    for cc in range(1, span + 1):
        db.cell(row=row, column=cc).fill = PatternFill("solid", fgColor=HEADER_BG)
    db.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    db.row_dimensions[row].height = 22


# --- the control: Excel re-derives the payment checks and the two must agree --
section_head(10, "CONTROL  -  EXCEL RE-DERIVES THE PAYMENT CHECKS AND IS COMPARED TO THE SQL RESULT")
db.cell(row=11, column=1,
        value="The Recon Detail tab recomputes the balance identity in Excel formulas from the raw remittance amounts, "
              "with no reference to what the SQL concluded. If the two ever part company, the last row stops reading zero.")
db.cell(row=11, column=1).font = Font(name=FONT, size=9, italic=True, color=MUTED)
db.merge_cells(start_row=11, start_column=1, end_row=11, end_column=8)

ctrl_head = ["Check", "What it tests", "SQL count", "Excel count", "Difference", "Status"]
for i, h in enumerate(ctrl_head, start=1):
    db.cell(row=13, column=i, value=h)
style_header(db, 13, len(ctrl_head))

CTRL = [
    ("DQ-05", "Claim marked PAID with no remittance line", "NO REMITTANCE (PAID)"),
    ("DQ-06", "Billed does not equal paid + patient responsibility + adjustment", "MISMATCH"),
    ("DQ-07", "Paid exceeds the billed charge", "OVERPAYMENT"),
]
for i, (cid, desc, verdict) in enumerate(CTRL):
    r = 14 + i
    db.cell(row=r, column=1, value=cid)
    db.cell(row=r, column=2, value=desc)
    db.cell(row=r, column=3, value=f'=COUNTIF({REG}!$B$2:$B${REG_LAST},"{cid}")').number_format = INT
    db.cell(row=r, column=4,
            value=f'=COUNTIF({RECON}!$S$2:$S${RECON_LAST},"{verdict}")').number_format = INT
    db.cell(row=r, column=5, value=f"=$C{r}-$D{r}").number_format = INT
    db.cell(row=r, column=6, value=f'=IF($E{r}=0,"AGREE","INVESTIGATE")')
    for c in range(1, 7):
        cell = db.cell(row=r, column=c)
        cell.font = Font(name=FONT, size=10, color=INK)
        cell.border = BOX

r = 17
db.cell(row=r, column=1, value="ALL")
db.cell(row=r, column=2, value="Claims where the Excel verdict and the SQL verdict disagree")
db.cell(row=r, column=3, value=f"={READ}!$C$12").number_format = INT
db.cell(row=r, column=4, value=f'=COUNTIF({RECON}!$U$2:$U${RECON_LAST},"AGREE")').number_format = INT
db.cell(row=r, column=5, value=f'=COUNTIF({RECON}!$U$2:$U${RECON_LAST},"CHECK")').number_format = INT
db.cell(row=r, column=6, value=f'=IF($E{r}=0,"AGREE","INVESTIGATE")')
for c in range(1, 7):
    cell = db.cell(row=r, column=c)
    cell.font = Font(name=FONT, size=10, bold=True, color=INK)
    cell.border = BOX

db.conditional_formatting.add("F14:F17",
    CellIsRule(operator="equal", formula=['"AGREE"'],
               fill=PatternFill("solid", fgColor=GOOD_BG),
               font=Font(name=FONT, size=10, bold=True, color=GOOD)))
db.conditional_formatting.add("F14:F17",
    CellIsRule(operator="equal", formula=['"INVESTIGATE"'],
               fill=PatternFill("solid", fgColor=BAD_BG),
               font=Font(name=FONT, size=10, bold=True, color=CRITICAL)))

# --- severity mix ---------------------------------------------------------
section_head(19, "EXCEPTIONS BY SEVERITY")
for i, h in enumerate(["Severity", "Exceptions", "Share of exceptions", "Financial exposure (USD)"], start=1):
    db.cell(row=20, column=i, value=h)
style_header(db, 20, 4)
for i, sev in enumerate(["CRITICAL", "HIGH", "MEDIUM"]):
    r = 21 + i
    db.cell(row=r, column=1, value=sev)
    db.cell(row=r, column=2, value=f'=COUNTIF({REG}!$C$2:$C${REG_LAST},"{sev}")').number_format = INT
    db.cell(row=r, column=3,
            value=f"=IF(COUNTA({REG}!$A$2:$A${REG_LAST})=0,0,$B{r}/COUNTA({REG}!$A$2:$A${REG_LAST}))").number_format = PCT1
    db.cell(row=r, column=4,
            value=f'=SUMIF({REG}!$C$2:$C${REG_LAST},"{sev}",{REG}!$M$2:$M${REG_LAST})').number_format = MONEY
    for c in range(1, 5):
        cell = db.cell(row=r, column=c)
        cell.font = Font(name=FONT, size=10, color=INK)
        cell.border = BOX
db.cell(row=24, column=1, value="TOTAL").font = Font(name=FONT, size=10, bold=True)
db.cell(row=24, column=2, value="=SUM(B21:B23)").number_format = INT
db.cell(row=24, column=4, value="=SUM(D21:D23)").number_format = MONEY
for c in (1, 2, 4):
    db.cell(row=24, column=c).font = Font(name=FONT, size=10, bold=True, color=INK)
    db.cell(row=24, column=c).border = Border(top=Side(style="double", color=INK))
db.conditional_formatting.add("A21:A23",
    CellIsRule(operator="equal", formula=['"CRITICAL"'],
               font=Font(name=FONT, size=10, bold=True, color=CRITICAL)))

# --- charts ---------------------------------------------------------------
section_head(26, "EXCEPTIONS AND EXPOSURE BY CHECK")

cats = Reference(cs, min_col=1, min_row=CS_FIRST, max_row=CS_LAST)

ch1 = BarChart()
ch1.type = "col"
ch1.style = None
ch1.title = "Exceptions raised by check"
ch1.y_axis.title = "Exceptions"
ch1.x_axis.title = None
ch1.legend = None                      # a single series names itself in the title
ch1.gapWidth = 40
data1 = Reference(cs, min_col=6, min_row=CS_FIRST - 1, max_row=CS_LAST)
ch1.add_data(data1, titles_from_data=True)
ch1.set_categories(cats)
ch1.series[0].graphicalProperties.solidFill = ACCENT
ch1.series[0].graphicalProperties.line.noFill = True
ch1.height, ch1.width = 7.5, 15
db.add_chart(ch1, "A28")

ch2 = BarChart()
ch2.type = "col"
ch2.style = None
ch2.title = "Financial exposure by check (USD)"
ch2.y_axis.title = "USD"
ch2.x_axis.title = None
ch2.legend = None
ch2.gapWidth = 40
data2 = Reference(cs, min_col=7, min_row=CS_FIRST - 1, max_row=CS_LAST)
ch2.add_data(data2, titles_from_data=True)
ch2.set_categories(cats)
ch2.series[0].graphicalProperties.solidFill = ACCENT
ch2.series[0].graphicalProperties.line.noFill = True
ch2.height, ch2.width = 7.5, 15
db.add_chart(ch2, "E28")

db.sheet_view.showGridLines = False

# Reading order: the summary first, the evidence behind it after.
ORDER = ["Dashboard", "Read Me", "Check Summary", "Defect Log",
         "Exception Register", "Recon Detail", "Payer Scorecard", "Suite Validation"]
wb._sheets = [wb[name] for name in ORDER]

for sheet in wb.worksheets:
    sheet.sheet_properties.tabColor = HEADER_BG if sheet.title == "Dashboard" else None

# Print setup: wide audit tabs are unreadable split across portrait pages.
for sheet in wb.worksheets:
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    if sheet.title in ("Exception Register", "Recon Detail"):
        sheet.print_title_rows = "1:1"
    elif sheet.title in ("Check Summary", "Defect Log", "Payer Scorecard", "Suite Validation"):
        sheet.print_title_rows = "4:4"

os.makedirs(OUT, exist_ok=True)
wb.save(XLSX)
print(f"written: {XLSX}")
print(f"  exception register rows : {REG_ROWS:,}")
print(f"  reconciliation rows     : {RECON_ROWS:,}")
print(f"  register spans          : {spans['DQ-05']}, {spans['DQ-06']}, {spans['DQ-07']}")
